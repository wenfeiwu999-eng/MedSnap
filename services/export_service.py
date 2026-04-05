# -*- coding: utf-8 -*-
"""Excel export service"""

import json, os
from datetime import datetime
import pandas as pd
from config import app, DEPARTMENT_CONFIGS

# ========== Excel 导出 ==========
def generate_excel(data_list):
    """多角色Excel导出，不同角色放不同Sheet"""
    from openpyxl.styles import Font, PatternFill
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red_font = Font(color='CC0000', bold=True)

    # 按角色分组
    grouped = {}
    for item in data_list:
        role = item.get('role_id', 'general')
        role_name = DEPARTMENT_CONFIGS.get(role, {}).get('name', role)
        grouped.setdefault(role_name, []).append(item)

    excel_path = os.path.join(
        app.config['UPLOAD_FOLDER'],
        f"临床数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_name, items in grouped.items():
            rows = []
            low_conf_fields = set()

            for item in items:
                row = {
                    '病历编号': item.get('case_number', ''),
                    '模板': item.get('template_name', ''),
                    '录入时间': item.get('create_time', ''),
                    '数据来源': '录音' if item.get('source_type') == 'audio' else '图片',
                }
                data = item.get('extracted_data', {})
                conf = item.get('confidence_data', {})

                # 递归展平JSON数据为Excel列
                _flatten_to_row(data, row, '', conf, low_conf_fields)

                # 音频数据额外列
                if item.get('source_type') == 'audio' and item.get('audio_transcript'):
                    transcript = item['audio_transcript']
                    row['转录原文'] = transcript[:500] + ('...' if len(transcript) > 500 else '')
                qual = item.get('qualitative_data')
                if qual and isinstance(qual, dict):
                    row['主题分析'] = ', '.join(qual.get('themes', []))
                    row['关键词'] = ', '.join(qual.get('keywords', []))
                    row['情感倾向'] = qual.get('sentiment', '')

                rows.append(row)

            if not rows:
                continue

            df = pd.DataFrame(rows)
            safe_name = sheet_name[:31]  # Excel sheet名最长31字符
            df.to_excel(writer, index=False, sheet_name=safe_name)

            # 标红低置信度
            if low_conf_fields:
                ws = writer.sheets[safe_name]
                for col_idx, col_name in enumerate(df.columns, 1):
                    for field_name in low_conf_fields:
                        if field_name in col_name:
                            for row_idx in range(2, len(df) + 2):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.fill = red_fill
                                cell.font = red_font
                            ws.cell(row=1, column=col_idx).fill = red_fill
                            ws.cell(row=1, column=col_idx).font = red_font
                            break

    return excel_path


def generate_unified_excel(data_list, batch_id=None):
    """统一数据集导出：所有记录合并到一个Sheet，两遍扫描确保列对齐"""
    from openpyxl.styles import Font, PatternFill
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    red_font = Font(color='CC0000', bold=True)

    tag = batch_id or datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = os.path.join(
        app.config['UPLOAD_FOLDER'],
        f"统一数据集_{tag}.xlsx"
    )

    all_rows = []
    low_conf_fields = set()

    for item in data_list:
        source_map = {'audio': '录音', 'text': '文本', 'image': '图片'}
        row = {
            '病历编号': item.get('case_number', ''),
            '原始文件': item.get('original_filename', ''),
            '模板': item.get('template_name', ''),
            '录入时间': item.get('create_time', ''),
            '数据来源': source_map.get(item.get('source_type', ''), item.get('source_type', '')),
        }
        data = item.get('extracted_data', {})
        conf = item.get('confidence_data', {})
        _flatten_to_row(data, row, '', conf, low_conf_fields)

        if item.get('source_type') == 'audio' and item.get('audio_transcript'):
            transcript = item['audio_transcript']
            row['转录原文'] = transcript[:500] + ('...' if len(transcript) > 500 else '')
        qual = item.get('qualitative_data')
        if qual and isinstance(qual, dict):
            row['主题分析'] = ', '.join(qual.get('themes', []))
            row['关键词'] = ', '.join(qual.get('keywords', []))
            row['情感倾向'] = qual.get('sentiment', '')

        all_rows.append(row)

    if not all_rows:
        return None

    df = pd.DataFrame(all_rows)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='统一数据集')

        if low_conf_fields:
            ws = writer.sheets['统一数据集']
            for col_idx, col_name in enumerate(df.columns, 1):
                for field_name in low_conf_fields:
                    if field_name in col_name:
                        for row_idx in range(2, len(df) + 2):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = red_fill
                            cell.font = red_font
                        ws.cell(row=1, column=col_idx).fill = red_fill
                        ws.cell(row=1, column=col_idx).font = red_font
                        break

    return excel_path


def _flatten_to_row(data, row, prefix, confidence, low_conf_fields):
    """递归展平嵌套JSON为Excel行的列"""
    if isinstance(data, dict):
        for k, v in data.items():
            if k == 'confidence':
                # 收集低置信度字段
                _collect_low_conf(v, low_conf_fields)
                continue
            full_key = f"{prefix}_{k}" if prefix else k
            if isinstance(v, dict):
                _flatten_to_row(v, row, full_key, confidence, low_conf_fields)
            elif isinstance(v, list):
                _flatten_list_to_row(v, row, full_key, confidence, low_conf_fields)
            else:
                row[full_key] = v
    elif isinstance(data, list):
        _flatten_list_to_row(data, row, prefix, confidence, low_conf_fields)


def _flatten_list_to_row(data_list, row, prefix, confidence, low_conf_fields):
    """展平列表数据"""
    for i, item in enumerate(data_list):
        if isinstance(item, dict):
            # 尝试用名称作为键
            name_key = item.get('项目名称') or item.get('英文缩写') or item.get('项目') or item.get('诊断名称') or str(i + 1)
            item_prefix = f"{prefix}_{name_key}" if prefix else name_key
            for k, v in item.items():
                if k in ('项目名称', '英文缩写', '项目', '诊断名称'):
                    continue
                if isinstance(v, (dict, list)):
                    continue
                col = f"{item_prefix}_{k}" if k != '数值' and k != '评分' else item_prefix
                row[col] = v
        elif isinstance(item, str):
            row[f"{prefix}_{i+1}"] = item


def _collect_low_conf(conf, low_conf_fields):
    """收集置信度<0.9的字段"""
    if isinstance(conf, dict):
        for k, v in conf.items():
            if isinstance(v, dict):
                _collect_low_conf(v, low_conf_fields)
            else:
                try:
                    if float(v) < 0.9:
                        low_conf_fields.add(k)
                except (ValueError, TypeError):
                    pass


# ========== 路由: 用户认证 ==========

